"""
cron-urlaubspreise - generalisierte, token-freie Flugpreis-Abfrage.

Liest die Konfiguration (config.xlsx, Ort steht in config-pfad.json), fragt fuer jede
aktive Reise alle Routen (Abflug x Ziel)
ueber mehrere Datums-Proben im Reisezeitraum per HEADLESS Browser (Playwright) bei
Google Flights ab, filtert nach Preis-Schwelle und max. Flugdauer, und schreibt:
  - best-of.xlsx   (Reiter "Best Of" global + je Reise ein Reiter)
  - lauf.log       (pro Reise ein Eintrag)

Aufruf:
  python flugpreise.py                 # alle aktiven Reisen, headless
  python flugpreise.py --sichtbar      # sichtbarer Browser (Debug)
  python flugpreise.py --reise kenia
  python flugpreise.py --limit 4       # max. N Routen-Abfragen gesamt (schneller Test)

Dauer-Filter beide Richtungen: Pro Route/Datum wird zusaetzlich die Rueck-Richtung
(dest->orig am Rueckdatum) als eigene Suche abgefragt. Eine Route/ein Datum gilt nur
als gueltig, wenn es in BEIDEN Richtungen Fluege <= max_flugdauer gibt. dauer_rueck =
kuerzeste gueltige Rueck-Richtungs-Dauer (Naeherung: bestaetigt, dass ein gueltiger
Rueckflug existiert; keine itinerar-genaue Kopplung an den jeweiligen Preis - das waere
nur ueber den fragilen Klick-Drill-in moeglich). Kostet ~2x Abfragen pro Route/Datum.
"""
import argparse
import base64
import datetime as dt
import json
import re
import sys
import time
from pathlib import Path

from playwright.sync_api import sync_playwright

BASE = Path(__file__).resolve().parent


def _pfade():
    """Zentrale Pfad-Konfiguration aus config-pfad.json im Projektordner:
    'zielordner' = Ordner fuer alle Dateien, 'config' und 'best_of_datei' =
    Dateinamen relativ zum Zielordner."""
    zeiger = BASE / "config-pfad.json"
    if zeiger.exists():
        return json.loads(zeiger.read_text(encoding="utf-8-sig"))
    return {}


def _abs(p, standard):
    """Pfad aufloesen: absolut wird direkt verwendet, relativ gilt relativ zu BASE."""
    p = p or standard
    return Path(p) if Path(p).is_absolute() else BASE / p


PFADE = _pfade()
ZIEL = _abs(PFADE.get("zielordner"), ".")
CONFIG = ZIEL / PFADE.get("config", "config.json")


def _ja(v):
    return str(v).strip().lower() in ("ja", "j", "x", "wahr", "true", "1")


def _liste(v):
    return [t.strip() for t in str(v or "").split(",") if t.strip()]


def _ziele_mit_aufenthalt(v):
    """Zerlegt 'SFO (72), HNL(48), MNL' in (['SFO','HNL','MNL'], [72, 48, None]).
    Die Klammer-Angabe ist der Mindestaufenthalt in Stunden (Hopping); ohne Klammer None."""
    codes, stunden = [], []
    for t in _liste(v):
        m = re.match(r"^(.*?)\s*\((\d+)\)$", t)
        if m:
            codes.append(re.sub(r"\s*/\s*", "/", m.group(1).strip()))
            stunden.append(int(m.group(2)))
        else:
            codes.append(re.sub(r"\s*/\s*", "/", t))
            stunden.append(None)
    return codes, stunden


def lade_config(pfad):
    """Laedt die Konfiguration aus config.xlsx (Reiter 'Reisen' + 'Einstellungen')
    und liefert dieselbe Struktur wie frueher die config.json. Ein .json-Pfad wird
    weiterhin direkt geladen (Fallback/Altbestand)."""
    if pfad.suffix.lower() == ".json":
        return json.loads(pfad.read_text(encoding="utf-8-sig"))

    from openpyxl import load_workbook
    # OneDrive/Excel koennen die Datei kurzzeitig sperren: ein paar Versuche mit Wartezeit
    for versuch in range(4):
        try:
            wb = load_workbook(pfad, data_only=True, read_only=True)
            break
        except PermissionError:
            if versuch == 3:
                raise
            time.sleep(5)

    einst = {}
    for row in wb["Einstellungen"].iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            einst[str(row[0]).strip()] = row[1]

    hopping_default_h = int(einst.get("hopping_min_aufenthalt_h", 48))

    konstellationen = []
    blatt = "Reisen" if "Reisen" in wb.sheetnames else "Konstellationen"  # alter Reitername als Fallback
    zeilen = wb[blatt].iter_rows(values_only=True)
    header = [str(h).strip().lower() if h is not None else "" for h in next(zeilen)]
    for row in zeilen:
        d = dict(zip(header, row))
        if not d.get("name"):
            continue
        ziele, aufenthalt = _ziele_mit_aufenthalt(d.get("ziele"))
        k = {
            "name": str(d["name"]).strip(),
            "aktiv": _ja(d.get("aktiv")),
            "typ": str(d.get("typ") or "return").strip().lower(),
            "ziele": ziele,
            "zeitraum": str(d.get("zeitraum") or "").strip(),
            "naechte": [int(x) for x in _liste(d.get("naechte"))],
        }
        if k["typ"] == "hopping":
            # Mindestaufenthalt je Station in Stunden; ohne Klammer-Angabe gilt der Default
            k["min_aufenthalt_h"] = [s if s is not None else hopping_default_h for s in aufenthalt]
        if d.get("schwelle_eur") is not None:
            k["schwelle_eur"] = float(d["schwelle_eur"])
        if d.get("personen"):
            k["personen"] = int(d["personen"])
        if d.get("max_flugdauer_h"):
            k["max_flugdauer_h_pro_richtung"] = float(d["max_flugdauer_h"])
        konstellationen.append(k)
    wb.close()

    return {
        "waehrung": str(einst.get("waehrung", "EUR")).strip(),
        "abflug": _liste(einst.get("abflug")),
        "job": {"zeitplan": {
            "frequenz_stunden": int(einst.get("frequenz_stunden", 1)),
            "verpasste_nachholen": _ja(einst.get("verpasste_nachholen")),
        }},
        "ausgaben": {
            "best_of_excel": {
                "datei": str(einst.get("best_of_datei", "best-of.xlsx")).strip(),
                "n": int(einst.get("n_best", 3)),
            },
            "log_datei": str(einst.get("log_datei", "lauf.log")).strip(),
        },
        "hopping_min_aufenthalt_h": hopping_default_h,
        "konstellationen": konstellationen,
    }

# --- bekannter, funktionierender tfs-Bauplan (FRA->MBA, 2026-08-08 / 2026-08-22) ---
# Zwei Legs mit FIXEN Feldlaengen: Datum (10), Flughafen (3). Daher Byte-Ersetzung moeglich.
_TEMPLATE_TFS = (
    "CBwQAhoeEgoyMDI2LTA4LTA4agcIARIDRlJBcgcIARIDTUJBGh4SCjIwMjYtMDgtMjJqBwgBEgN"
    "NQkFyBwgBEgNGUkFAAUgBcAGCAQsI____________AZgBAQ")
_TEMPLATE = base64.urlsafe_b64decode(_TEMPLATE_TFS + "=" * (-len(_TEMPLATE_TFS) % 4))
# Byte-Positionen (aus Dekodierung verifiziert):
_POS = {"dep_date": 8, "out_orig": 24, "out_dest": 33,
        "ret_date": 40, "ret_orig": 56, "ret_dest": 65}

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")

# Soft-Deadline pro Reise: bricht kontrolliert ab, BEVOR die Windows-Aufgabe
# den Lauf nach 30 Min hart killt. So werden Teilergebnis, Marker-Ersetzung und State
# noch sauber geschrieben (statt mitten im Lauf verloren zu gehen).
SOFT_DEADLINE_S = 24 * 60


def append_log(log_path, line):
    with log_path.open("a", encoding="utf-8") as f:
        f.write(line + "\n")
        f.flush()
    print(line)


def tag_trenner_falls_noetig(log_path):
    """Setzt zwei Strich-Zeilen vor den ersten Eintrag eines neuen Tages,
    um Tageswechsel im Log optisch klar abzusetzen."""
    if not log_path.exists():
        return
    zeilen = [z for z in log_path.read_text(encoding="utf-8").splitlines() if z.strip()]
    if not zeilen:
        return
    if set(zeilen[-1].strip()) == {"-"}:
        return  # Log endet bereits mit einem Trenner
    letztes_datum = None
    for z in reversed(zeilen):
        if set(z.strip()) == {"-"}:
            continue
        letztes_datum = z[:10]  # fuehrendes YYYY-MM-DD des letzten Eintrags
        break
    if letztes_datum == dt.date.today().isoformat():
        return  # heute wurde bereits etwas geloggt -> kein neuer Tag
    with log_path.open("a", encoding="utf-8") as f:
        f.write(("-" * 80 + "\n") * 2)


def bereinige_marker(log_path):
    """Entfernt verwaiste 'LAEUFT'-Zeilen frueherer Laeufe, die abgewuergt wurden
    (z. B. PC-Standby oder Zeitlimit), bevor sie ihre finale Zeile schreiben konnten.
    Da immer nur eine Instanz laeuft, ist jede beim Start vorhandene LAEUFT-Zeile verwaist."""
    if not log_path.exists():
        return
    zeilen = log_path.read_text(encoding="utf-8").splitlines(keepends=True)
    behalten = [z for z in zeilen if "| Status: LAEUFT |" not in z]
    if len(behalten) != len(zeilen):
        log_path.write_text("".join(behalten), encoding="utf-8")


def rolliere_log(log_path, tage=7):
    """Haelt das Log rollierend auf den gleichen Zeitraum wie die Historie in Excel/App:
    Eintraege aelter als 'tage' Tage werden entfernt (das Log ist chronologisch, also
    wird einfach der alte Anfang abgeschnitten, inklusive verwaister Tages-Trenner)."""
    if not log_path.exists():
        return
    grenze = (dt.date.today() - dt.timedelta(days=tage - 1)).isoformat()
    text = log_path.read_text(encoding="utf-8", errors="replace")
    sauber = text.replace("\x00", "")  # NUL-Reste aus frueheren Schreib-Kollisionen tilgen
    zeilen = [z for z in sauber.splitlines(keepends=True) if z.strip("\n\r ")]
    idx = None
    for i, z in enumerate(zeilen):
        d = z[:10]
        if len(d) == 10 and d[4:5] == "-" and d >= grenze:
            idx = i
            break
    neu = "".join(zeilen[idx:]) if idx is not None else ""
    if neu != text:
        log_path.write_text(neu, encoding="utf-8")


def schreibe_lauf_marker(log_path, line):
    """Schreibt eine provisorische 'LAEUFT'-Zeile ans Ende und liefert die Byte-Position
    VOR der Zeile zurueck, damit sie spaeter exakt ersetzt/entfernt werden kann."""
    pos = log_path.stat().st_size if log_path.exists() else 0
    append_log(log_path, line)
    return pos


def ersetze_lauf_marker(log_path, pos, line=None):
    """Ersetzt die zuvor geschriebene LAEUFT-Markerzeile durch die finale Zeile (oder
    entfernt sie, wenn line None ist). Sucht die Markerzeile im aktuellen Inhalt statt
    blind an einer Byte-Position zu schreiben - robust gegen zwischenzeitliche
    Log-Aenderungen durch parallele Laeufe (frueher entstanden dadurch NUL-Luecken)."""
    zeilen = []
    if log_path.exists():
        zeilen = log_path.read_text(encoding="utf-8").splitlines(keepends=True)
    marker = "| Status: LAEUFT |"
    if line is not None and line.count(" | ") >= 2:
        # eigenen Marker anhand des Reise-Namens finden (parallele Laeufe moeglich)
        marker = "| " + line.split(" | ")[1] + " " + marker
    idx = next((i for i in range(len(zeilen) - 1, -1, -1)
                if marker in zeilen[i]), None)
    if idx is None:
        idx = next((i for i in range(len(zeilen) - 1, -1, -1)
                    if "| Status: LAEUFT |" in zeilen[i]), None)
    if idx is not None:
        del zeilen[idx]
    if line is not None:
        zeilen.append(line + "\n")
    log_path.write_text("".join(zeilen), encoding="utf-8")
    if line is not None:
        print(line)


def build_tfs(orig: str, dest: str, dep: str, ret: str) -> str:
    """Baut den tfs-Parameter fuer eine Route (orig->dest) mit Hin-/Rueckdatum (YYYY-MM-DD)."""
    b = bytearray(_TEMPLATE)
    def put(pos, s, n):
        b[pos:pos + n] = s.encode()
    put(_POS["dep_date"], dep, 10)
    put(_POS["out_orig"], orig, 3)
    put(_POS["out_dest"], dest, 3)
    put(_POS["ret_date"], ret, 10)
    put(_POS["ret_orig"], dest, 3)
    put(_POS["ret_dest"], orig, 3)
    return base64.urlsafe_b64encode(bytes(b)).decode().rstrip("=")


def search_url(orig, dest, dep, ret):
    return "https://www.google.com/travel/flights/search?tfs=" + build_tfs(orig, dest, dep, ret)


def build_tfs_legs(legs, trip_typ):
    """Baut den tfs-Parameter fuer beliebige Etappen [(orig, dest, dep), ...].
    trip_typ: 1 = Hin+Rueck, 2 = One-Way, 3 = Multi-City. Der Etappen-Block (Bytes 4-35)
    und das Trip-Typ-Byte (letztes Byte, Feld 19) stammen aus dem verifizierten Template."""
    teile = [_TEMPLATE[:4]]
    for orig, dest, dep in legs:
        leg = bytearray(_TEMPLATE[4:36])
        leg[4:14] = dep.encode()
        leg[20:23] = orig.encode()
        leg[29:32] = dest.encode()
        teile.append(bytes(leg))
    teile.append(_TEMPLATE[68:-1] + bytes([trip_typ]))
    return base64.urlsafe_b64encode(b"".join(teile)).decode().rstrip("=")


def search_url_oneway(orig, dest, dep):
    return "https://www.google.com/travel/flights/search?tfs=" + build_tfs_legs([(orig, dest, dep)], 2)


def search_url_kette(legs):
    """Multi-City-Suche mit allen Etappen vorbefuellt - als Buchungs-Link je Ketten-Angebot."""
    return "https://www.google.com/travel/flights/search?tfs=" + build_tfs_legs(legs, 3)


def parse_zeitraum(s: str):
    """'08.08.2026-28.08.2026' -> (date, date)."""
    a, b = s.split("-")
    f = lambda x: dt.datetime.strptime(x.strip(), "%d.%m.%Y").date()
    return f(a), f(b)


def datums_proben(zeitraum, naechte, schritt_tage=4):
    """Erzeugt (dep, ret)-Paare im Fenster: dep gestuft, naechte = [min, max]."""
    start, ende = parse_zeitraum(zeitraum)
    nmin, nmax = naechte
    paare = []
    dep = start
    while dep <= ende:
        for n in sorted({nmin, nmax}):
            ret = dep + dt.timedelta(days=n)
            if ret <= ende:
                paare.append((dep.isoformat(), ret.isoformat()))
        dep += dt.timedelta(days=schritt_tage)
    return paare


_DUR = re.compile(r"(\d+)\s*Std\.?(?:\s*(\d+)\s*Min)?")  # "17 Std." (glatt) ODER "14 Std. 45 Min."
_PRICE = re.compile(r"(\d[\d.]{2,5})\s?€")
_STOPS = re.compile(r"(\d+)\s*Stopp")  # "1 Stopp" / "2 Stopps"; Nonstop = 0


def dur_to_min(text):
    """Groesste 'X Std. Y Min.'-Angabe = Gesamt-Bruttodauer (Layover ist immer kuerzer)."""
    best = None
    for m in _DUR.finditer(text):
        h = int(m.group(1)); mi = int(m.group(2) or 0)
        val = h * 60 + mi
        if best is None or val > best:
            best = val
    return best


def handle_consent(page):
    if "consent" in page.url or page.locator("text=Bevor Sie fortfahren").count() > 0:
        for label in ["Alle ablehnen", "Reject all", "Alle akzeptieren", "Accept all"]:
            btn = page.get_by_role("button", name=label)
            if btn.count() > 0:
                btn.first.click()
                page.wait_for_timeout(2500)
                return True
    return False


def query_route(page, orig, dest, dep, ret):
    """Liefert Liste von Angeboten [{preis, dauer_hin_min, airline}] fuer eine Route+Datum."""
    url = search_url(orig, dest, dep, ret)
    geladen = False
    for attempt in range(2):  # ein Retry bei traegem Laden
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(2500)
        handle_consent(page)
        try:
            page.wait_for_selector("text=/\\d[\\d.]*\\s?€/", timeout=20000)
            geladen = True
            break
        except Exception:
            page.wait_for_timeout(1500)
    if not geladen:
        return None  # nichts geladen / blockiert
    page.wait_for_timeout(1500)

    angebote = []
    # Flug-Ergebnisse sind Listeneintraege; aria-label enthaelt Dauer+Preis+Airline.
    items = page.locator("li").all()
    for it in items:
        try:
            txt = (it.get_attribute("aria-label") or "") + " " + it.inner_text()
        except Exception:
            continue
        # Nur echte Round-Trip-Ergebniskarten (Marker "Hin und zur[ueck]") akzeptieren -
        # schliesst Fragmente/Teilstrecken aus, die unmoegliche Mini-Dauern (z.B. 2h) liefern.
        if "€" not in txt or "Std" not in txt or "Hin und zur" not in txt:
            continue
        pm = _PRICE.search(txt)
        dh = dur_to_min(txt)
        if not pm or dh is None:
            continue
        preis = int(pm.group(1).replace(".", ""))
        if not (100 <= preis <= 20000):
            continue
        air = ""
        # Airline = Zeile direkt vor der ersten Gesamtdauer-Zeile ("X Std.").
        # In der Ergebniskarte steht der Carrier (ggf. mehrere bei Codeshare) genau dort,
        # nach den Uhrzeiten und vor der Dauer. Generisch statt fester Carrier-Liste.
        zeilen = [l.strip() for l in txt.splitlines() if l.strip()]
        for i, l in enumerate(zeilen):
            if re.match(r"^\d+\s*Std", l):
                if i > 0 and not any(ch.isdigit() for ch in zeilen[i - 1]):
                    air = zeilen[i - 1][:60]
                break
        sm = _STOPS.search(txt)
        umstiege = int(sm.group(1)) if sm else (0 if "Nonstop" in txt else None)
        angebote.append({"preis": preis, "dauer_hin_min": dh, "airline": air, "umstiege": umstiege})
    # dedupe + sort
    seen, out = set(), []
    for a in sorted(angebote, key=lambda x: x["preis"]):
        key = (a["preis"], a["dauer_hin_min"], a["umstiege"])
        if key not in seen:
            seen.add(key); out.append(a)
    return out


def query_leg(page, orig, dest, dep):
    """Liefert One-Way-Angebote [{preis, dauer_hin_min, airline, umstiege}] fuer eine Etappe."""
    url = search_url_oneway(orig, dest, dep)
    geladen = False
    for attempt in range(2):
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(2500)
        handle_consent(page)
        try:
            page.wait_for_selector("text=/\\d[\\d.]*\\s?€/", timeout=20000)
            geladen = True
            break
        except Exception:
            page.wait_for_timeout(1500)
    if not geladen:
        return None
    page.wait_for_timeout(1500)

    angebote = []
    for it in page.locator("li").all():
        try:
            txt = (it.get_attribute("aria-label") or "") + " " + it.inner_text()
        except Exception:
            continue
        # Echte Ergebniskarten enthalten beide Flughafencodes (z. B. "FRA–SFO");
        # das schliesst Vorschlags-/Fragment-Zeilen aus.
        if "€" not in txt or "Std" not in txt or orig not in txt or dest not in txt:
            continue
        pm = _PRICE.search(txt)
        dh = dur_to_min(txt)
        if not pm or dh is None:
            continue
        preis = int(pm.group(1).replace(".", ""))
        if not (30 <= preis <= 20000):  # One-Way-Etappen koennen guenstig sein
            continue
        air = ""
        zeilen = [l.strip() for l in txt.splitlines() if l.strip()]
        for i, l in enumerate(zeilen):
            if re.match(r"^\d+\s*Std", l):
                if i > 0 and not any(ch.isdigit() for ch in zeilen[i - 1]):
                    air = zeilen[i - 1][:60]
                break
        sm = _STOPS.search(txt)
        umstiege = int(sm.group(1)) if sm else (0 if "Nonstop" in txt else None)
        angebote.append({"preis": preis, "dauer_hin_min": dh, "airline": air, "umstiege": umstiege})
    seen, out = set(), []
    for a in sorted(angebote, key=lambda x: x["preis"]):
        key = (a["preis"], a["dauer_hin_min"], a["umstiege"])
        if key not in seen:
            seen.add(key); out.append(a)
    return out


def hopping_proben(k, schritt_tage):
    """Startdaten fuer eine Hopping-Kette: gestuft im Zeitraum, sodass die Kette
    (Flugtage + Mindestaufenthalte) noch komplett ins Fenster passt. Max. 4 Proben."""
    start, ende = parse_zeitraum(k["zeitraum"])
    aufenthalt = k.get("min_aufenthalt_h", [])
    grob = sum(1 + -(-h // 24) for h in aufenthalt) + 2  # je Station Flugtag + Aufenthalt, + Rueckflug
    if k.get("naechte"):
        grob = max(grob, min(k["naechte"]))  # Mindest-Gesamtdauer der Reise
    proben = []
    d = start
    while d + dt.timedelta(days=grob) <= ende and len(proben) < 4:
        proben.append(d)
        d += dt.timedelta(days=max(schritt_tage, 2))
    return proben


def bepreise_kette(page, k, abf, startdatum, maxdauer_min, schwelle, cache):
    """Bepreist alle Varianten einer Hopping-Kette ab 'abf': Alternativen je Station
    stehen per '/' im Ziel (z. B. 'SFO/LAX'); jede Kombination ist eine eigene Kette.
    Liefert (dict variante->angebot, neue_abfragen, blockierte)."""
    import itertools
    alternativen = [z.split("/") for z in k["ziele"]]
    kombis = list(itertools.product(*alternativen))
    if len(kombis) > 12:
        print(f"[!] {k['name']}: {len(kombis)} Ketten-Varianten - nur die ersten 12 werden bepreist.")
        kombis = kombis[:12]
    ergebnisse = {}
    n_gesamt, blockiert_gesamt = 0, 0
    for stationen in kombis:
        kette, n, b = _bepreise_feste_kette(page, k, abf, startdatum, list(stationen),
                                            maxdauer_min, schwelle, cache)
        n_gesamt += n
        blockiert_gesamt += b
        if kette:
            ergebnisse[kette["dest"]] = kette
    return ergebnisse, n_gesamt, blockiert_gesamt


def _bepreise_feste_kette(page, k, abf, startdatum, stationen, maxdauer_min, schwelle, cache):
    """Bepreist EINE konkrete Kette (feste Stationsfolge): jede Etappe als One-Way-Suche,
    Folgetermine aus Ankunftstag + Mindestaufenthalt (tagesgenau, +6 h Puffer fuer
    Abflugzeit/Zeitzonen). Liefert (angebot|None, neue_abfragen, blockierte)."""
    aufenthalt = k.get("min_aufenthalt_h", [48] * len(stationen))
    naechte = k.get("naechte") or []
    nmin = min(naechte) if naechte else None
    nmax = max(naechte) if naechte else None
    _, zeitraum_ende = parse_zeitraum(k["zeitraum"])
    legs = [(abf, stationen[0])] + list(zip(stationen, stationen[1:])) + [(stationen[-1], abf)]
    dep = startdatum
    etappen, geplant = [], []
    n_abfragen = 0
    ende = None
    for i, (von, nach) in enumerate(legs):
        if i == len(stationen) and nmin is not None and (dep - startdatum).days < nmin - 1:
            # Kette kuerzer als Mindest-Gesamtdauer: Aufenthalt an der letzten Station
            # verlaengern, sodass die Rueckkehr fruehestens nach nmin Naechten liegt.
            dep = startdatum + dt.timedelta(days=nmin - 1)
        datum = dep.isoformat()
        key = (von, nach, datum)
        if key in cache:
            angebote = cache[key]
        else:
            angebote = query_leg(page, von, nach, datum)
            n_abfragen += 1
            cache[key] = angebote
        if angebote is None:
            return None, n_abfragen, 1
        gueltig = [a for a in angebote if a["dauer_hin_min"] <= maxdauer_min]
        if not gueltig:
            return None, n_abfragen, 0
        best = min(gueltig, key=lambda a: a["preis"])
        etappen.append({"von": von, "nach": nach, "datum": datum, "preis": best["preis"],
                        "dauer_min": best["dauer_hin_min"], "umstiege": best["umstiege"],
                        "airline": best["airline"]})
        geplant.append((von, nach, datum))
        ankunft = dep + dt.timedelta(days=-(-(best["dauer_hin_min"] + 360) // 1440))
        if i < len(stationen):
            dep = ankunft + dt.timedelta(days=-(-aufenthalt[i] // 24))
        else:
            ende = ankunft
    if ende > zeitraum_ende:
        return None, n_abfragen, 0  # Kette passt nicht mehr ins Reisefenster
    if nmax is not None and (ende - startdatum).days > nmax:
        return None, n_abfragen, 0  # Gesamtdauer ueber dem Maximum aus 'Naechte'
    preis = sum(e["preis"] for e in etappen)
    airlines = ", ".join(dict.fromkeys(e["airline"] for e in etappen if e["airline"]))
    return ({
        "orig": abf, "dest": "-".join(stationen), "preis": preis,
        "dauer_hin_min": etappen[0]["dauer_min"], "dauer_rueck_min": etappen[-1]["dauer_min"],
        "airline": airlines[:60],
        "umstiege_hin": etappen[0]["umstiege"], "umstiege_rueck": etappen[-1]["umstiege"],
        "reisetage": (ende - startdatum).days,
        "dep": startdatum.isoformat(), "ret": ende.isoformat(),
        "unter": (schwelle is not None and preis < schwelle),
        "etappen": etappen,
        "url": search_url_kette(geplant),
    }, n_abfragen, 0)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sichtbar", action="store_true", help="Browser sichtbar (Debug)")
    ap.add_argument("--reise", "--konstellation", dest="reise", help="nur diese Reise")
    ap.add_argument("--limit", type=int, default=0, help="max. Routen-Abfragen gesamt")
    ap.add_argument("--schritt-tage", type=int, default=4, help="Abstand der Datums-Proben")
    args = ap.parse_args()

    try:
        cfg = lade_config(CONFIG)
    except PermissionError:
        # Config dauerhaft gesperrt (z. B. in Excel geoeffnet): sauber protokollieren
        # statt abstuerzen; der naechste stuendliche Lauf versucht es erneut.
        zeitstempel = dt.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        append_log(BASE / "lauf.log",
                   f"{zeitstempel} | - | Status: CONFIG_GESPERRT | {CONFIG.name} ist gesperrt "
                   f"(in Excel geoeffnet?) - Lauf uebersprungen.")
        return
    abflug = cfg["abflug"]
    schwelle_default = None
    n_best = cfg["ausgaben"]["best_of_excel"].get("n", 3)
    log_path = BASE / cfg["ausgaben"].get("log_datei", "lauf.log")
    bereinige_marker(log_path)  # verwaiste LAEUFT-Zeilen abgewuergter Laeufe entfernen
    rolliere_log(log_path)      # Log rollierend halten (gleicher Zeitraum wie die Historie)
    # Pfad der best-of.xlsx: Zielordner + Dateiname aus config-pfad.json;
    # Fallback: Wert aus der Config (Altbestand) bzw. "best-of.xlsx".
    xlsx_path = ZIEL / PFADE.get("best_of_datei", cfg["ausgaben"]["best_of_excel"].get("datei", "best-of.xlsx"))
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    html_path = ZIEL / PFADE.get("best_of_html", "best-of.html")
    app_path = ZIEL / PFADE.get("best_of_app", "best-of-app.html")
    # Streunende Kopie am alten Standard-Ort entfernen, falls die Config woanders hinzeigt
    alt_pfad = BASE / "best-of.xlsx"
    if alt_pfad != xlsx_path and alt_pfad.exists():
        try:
            alt_pfad.unlink()
        except PermissionError:
            pass
    data_path = BASE / "best-of-data.json"
    hist_path = BASE / "historie.json"

    alle_konst = [k for k in cfg["konstellationen"] if k.get("aktiv")]

    if args.reise:
        konst = [k for k in alle_konst if k["name"] == args.reise]
        state_path = heute = naechster = None
    else:
        state_path = BASE / "state.json"
        heute = dt.date.today().isoformat()
        if state_path.exists():
            state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        else:
            state = {"datum": "", "letzter_index": -1, "fehlversuche": {}}

        gleicher_tag = state.get("datum") == heute
        fehlversuche = state.get("fehlversuche", {}) if gleicher_tag else {}
        resume = state.get("resume", {}) if gleicher_tag else {}

        if gleicher_tag and 0 <= state["letzter_index"] < len(alle_konst):
            name_letzter = alle_konst[state["letzter_index"]]["name"]
            if resume.get(name_letzter, 0) > 0:
                # letzter Lauf nur teilweise geschafft (Soft-Zeitlimit) -> hier weitermachen
                naechster = state["letzter_index"]
            elif 0 < fehlversuche.get(name_letzter, 0) < 2:
                # letzter Lauf fehlgeschlagen, aber noch ein Versuch übrig -> wiederholen
                naechster = state["letzter_index"]
            else:
                # vorwärts, überspringe ausgeschlossene (2 Fehlversuche)
                naechster = state["letzter_index"] + 1
                while naechster < len(alle_konst) and fehlversuche.get(alle_konst[naechster]["name"], 0) >= 2:
                    naechster += 1
                if naechster >= len(alle_konst):
                    tag_trenner_falls_noetig(log_path)
                    msg = f"{dt.datetime.now().strftime('%Y-%m-%dT%H:%M:%S')} | - | Status: NICHTS_ZU_TUN | Alle {len(alle_konst)} Reisen heute bereits abgearbeitet."
                    append_log(log_path, msg)
                    return
        else:
            naechster = 0

        konst = [alle_konst[naechster]]

    zeit = lambda: dt.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    best_pro_konst = {}   # name -> {route: bestes angebot}
    start_offsets = {}    # name -> Resume-Offset, mit dem dieser Lauf gestartet ist
    gesamt_abfragen = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=not args.sichtbar)
        ctx = browser.new_context(user_agent=UA, locale="de-DE",
                                  timezone_id="Europe/Berlin",
                                  viewport={"width": 1366, "height": 900})
        page = ctx.new_page()

        for k in konst:
            name = k["name"]
            ziele = k["ziele"]
            personen = k.get("personen", 1)
            schwelle = k.get("schwelle_eur")
            maxdauer_min = k.get("max_flugdauer_h_pro_richtung", 99) * 60

            # Provisorische 'LAEUFT'-Zeile: macht einen gerade laufenden (oder haengenden)
            # Lauf sofort im Log sichtbar. Wird am Ende durch die finale Zeile ersetzt.
            tag_trenner_falls_noetig(log_path)
            marker_pos = schreibe_lauf_marker(
                log_path, f"{zeit()} | {name} | Status: LAEUFT | Abfrage laeuft gerade ...")

            t0 = time.time()
            routen_gezaehlt = 0
            gueltige = 0
            blockiert = 0
            abgebrochen = False
            best_pro_konst[name] = {}
            status = "OK"
            hinweis = ""

            # Flache, deterministische Liste aller Abfragen.
            # return: (Abflug x Ziel x Datumsprobe); hopping: (Startdatum x Abflug).
            # Resume: bei nicht --reise ab gespeichertem Offset weitermachen.
            if k["typ"] == "hopping":
                combos = [(abf, probe) for probe in hopping_proben(k, args.schritt_tage)
                          for abf in abflug]
            else:
                naechte = k.get("naechte", [7, 7])
                proben = datums_proben(k["zeitraum"], naechte, args.schritt_tage)
                combos = [(orig, dest, dep, ret)
                          for orig in abflug for dest in ziele for dep, ret in proben]
            start = resume.get(name, 0) if (heute is not None) else 0
            if start >= len(combos):
                start = 0  # Sicherheitsnetz (z. B. Config verkleinert)
            start_offsets[name] = start
            verarbeitet = 0

            leg_cache = {}  # Hopping: (von, nach, datum) -> Angebote; Ketten teilen sich Etappen-Abfragen
            try:
                for combo in combos[start:]:
                    if time.time() - t0 > SOFT_DEADLINE_S:
                        abgebrochen = True
                        break
                    if args.limit and gesamt_abfragen >= args.limit:
                        break
                    routen_gezaehlt += 1
                    verarbeitet += 1
                    if k["typ"] == "hopping":
                        abf, probe = combo
                        ketten, n_neu, n_blockiert = bepreise_kette(
                            page, k, abf, probe, maxdauer_min, schwelle, leg_cache)
                        gesamt_abfragen += n_neu
                        blockiert += n_blockiert
                        for kv in ketten.values():
                            gueltige += 1
                            route = f"{abf}->{kv['dest']}"
                            cur = best_pro_konst[name].get(route)
                            if cur is None or kv["preis"] < cur["preis"]:
                                best_pro_konst[name][route] = kv
                        continue
                    orig, dest, dep, ret = combo
                    gesamt_abfragen += 1
                    route = f"{orig}->{dest}"
                    angebote = query_route(page, orig, dest, dep, ret)
                    if angebote is None:
                        blockiert += 1
                        continue
                    # Rueck-Richtung als eigene Suche (dest->orig am Rueckdatum):
                    # liefert zuverlaessig die Rueckflug-Dauern, ohne fragiles Klicken.
                    rueck = query_route(page, dest, orig, ret, dep)
                    if rueck is None:
                        blockiert += 1
                        continue
                    rueck_valid = [(r["dauer_hin_min"], r["umstiege"]) for r in rueck
                                   if r["dauer_hin_min"] <= maxdauer_min]
                    if not rueck_valid:
                        continue  # keine gueltige Rueckrichtung -> Route/Datum ungueltig
                    best_rueck_item = min(rueck_valid, key=lambda x: x[0])
                    best_rueck, umstiege_rueck = best_rueck_item
                    naechte_real = (dt.date.fromisoformat(ret) - dt.date.fromisoformat(dep)).days
                    for a in angebote:
                        if a["dauer_hin_min"] > maxdauer_min:
                            continue  # Dauer-Filter (Hinflug): ungueltig
                        unter = (schwelle is not None and a["preis"] < schwelle)
                        gueltige += 1
                        # bestes je Route merken
                        cur = best_pro_konst[name].get(route)
                        if cur is None or a["preis"] < cur["preis"]:
                            best_pro_konst[name][route] = {
                                "orig": orig, "dest": dest, "preis": a["preis"],
                                "dauer_hin_min": a["dauer_hin_min"],
                                "dauer_rueck_min": best_rueck, "airline": a["airline"],
                                "umstiege_hin": a["umstiege"], "umstiege_rueck": umstiege_rueck,
                                "reisetage": naechte_real,
                                "dep": dep, "ret": ret, "unter": unter}
            except Exception as e:
                status = "FEHLER"
                # nur erste Zeile, ohne mehrzeiligen Playwright-Call-Log
                hinweis = str(e).splitlines()[0][:120] if str(e) else ""

            dauer_s = int(time.time() - t0)
            bestp = min((v["preis"] for v in best_pro_konst[name].values()), default=None)
            bestr = None
            if bestp is not None:
                bestr = next(r for r, v in best_pro_konst[name].items() if v["preis"] == bestp)
            if status != "FEHLER":
                if blockiert and gueltige == 0:
                    status = "BLOCKIERT"
                    hinweis = f"{blockiert} Abfragen ohne Ergebnis (evtl. Captcha)"
                elif gueltige == 0:
                    status = "KEINE_TREFFER"
                elif blockiert:
                    status = "OK"
                    hinweis = f"{blockiert} von {routen_gezaehlt} Abfragen ohne Ergebnis"
            neuer_offset = start + verarbeitet
            fertig = neuer_offset >= len(combos)
            if abgebrochen and status != "FEHLER" and verarbeitet > 0 and not fertig:
                # Soft-Deadline gegriffen: Teilergebnis gesichert, der Rest wird im
                # naechsten Lauf ab 'neuer_offset' fortgesetzt (echtes Resume).
                status = "TEILWEISE"
                hinweis = (f"Soft-Zeitlimit {SOFT_DEADLINE_S // 60} Min erreicht - "
                           f"Abfrage {neuer_offset}/{len(combos)} erreicht, Rest wird fortgesetzt")
            # Nur loggen, wenn die Reise tatsaechlich abgefragt wurde.
            # Nicht behandelte (z. B. wegen --limit uebersprungene) erzeugen KEINE Zeile.
            # Die zu Beginn geschriebene 'LAEUFT'-Markerzeile wird hier ersetzt (bzw. entfernt).
            if routen_gezaehlt > 0:
                ersetze_lauf_marker(log_path, marker_pos,
                    f"{zeit()} | {name} | Status: {status} | Routen: {routen_gezaehlt} | "
                    f"gueltige Treffer: {gueltige} | guenstigster: "
                    f"{str(bestp)+' EUR ('+bestr+')' if bestp else '-'} | Dauer: {dauer_s}s | {hinweis}")
                if not args.reise and heute is not None:
                    if status == "TEILWEISE":
                        # Fortschritt, kein Fehler: Offset merken, Reise NICHT abhaken,
                        # Fehlerzaehler unveraendert lassen -> naechster Lauf macht hier weiter.
                        resume[name] = neuer_offset
                    else:
                        resume.pop(name, None)  # Voll-Durchlauf abgeschlossen
                        if status != "OK":
                            fehlversuche[name] = fehlversuche.get(name, 0) + 1
                        else:
                            fehlversuche.pop(name, None)  # Erfolg loescht den Fehlerzaehler
                    state_path.write_text(json.dumps({
                        "datum": heute, "letzter_index": naechster,
                        "fehlversuche": fehlversuche, "resume": resume
                    }), encoding="utf-8")
            else:
                # Reise wurde gar nicht abgefragt -> Markerzeile wieder entfernen.
                ersetze_lauf_marker(log_path, marker_pos, None)

        browser.close()

    # Akkumulierten Stand laden, aktuelle Reise einmergen, Excel neu schreiben.
    # Nur Reisen mit tatsaechlichen Ergebnissen ueberschreiben (leere Laeufe behalten alte Werte).
    if data_path.exists():
        gesamt = json.loads(data_path.read_text(encoding="utf-8"))
    else:
        gesamt = {}
    for nm, routen in best_pro_konst.items():
        if not routen:
            continue
        if start_offsets.get(nm, 0) == 0:
            gesamt[nm] = routen   # frischer Voll-Durchlauf -> komplett ersetzen
        else:
            # Resume-Lauf: nur diesen Teil einmergen, je Route den besten Preis behalten.
            ziel = gesamt.setdefault(nm, {})
            for rt, off in routen.items():
                if rt not in ziel or off["preis"] < ziel[rt]["preis"]:
                    ziel[rt] = off
    # Aus der Config entfernte/deaktivierte Reisen aus dem Akku-Speicher werfen,
    # sonst bleiben sie als "Leichen" im Excel stehen und bekommen bei jedem Lauf nur einen
    # frischen "Stand"-Zeitstempel (obwohl ihre Preise eingefroren sind).
    aktive_namen = {k["name"] for k in alle_konst}
    for nm in [n for n in gesamt if n not in aktive_namen]:
        del gesamt[nm]
    data_path.write_text(json.dumps(gesamt, ensure_ascii=False, indent=2), encoding="utf-8")
    historie = update_historie(hist_path, best_pro_konst)
    schreibe_excel(xlsx_path, gesamt, n_best, historie)
    schwellen = {k["name"]: k.get("schwelle_eur") for k in alle_konst}
    schreibe_html(html_path, gesamt, n_best, historie, schwellen)
    schreibe_app_html(app_path, gesamt, n_best, historie, schwellen)
    if PFADE.get("github_pages_repo"):
        veroeffentliche_html(app_path, PFADE["github_pages_repo"])
    print(f"Fertig. Abfragen: {gesamt_abfragen}.")


def update_historie(hist_path, best_pro_konst, tage=7):
    """Pflegt eine rollierende 7-Tage-Historie: pro (Datum, Reise, Verbindung)
    der beste Preis des Tages. Eintraege aelter als 'tage' Tage werden verworfen."""
    heute = dt.date.today().isoformat()
    if hist_path.exists():
        hist = json.loads(hist_path.read_text(encoding="utf-8-sig"))
    else:
        hist = []
    idx = {(r["datum"], r["konstellation"], r["route"]): r for r in hist}
    for name, routen in best_pro_konst.items():
        if not routen:
            continue
        for route, v in routen.items():
            rec = {
                "datum": heute, "konstellation": name, "route": route,
                "orig": v["orig"], "dest": v["dest"], "preis": v["preis"],
                "dep": v["dep"], "ret": v["ret"], "airline": v.get("airline", ""),
                "umstiege_hin": v.get("umstiege_hin"), "umstiege_rueck": v.get("umstiege_rueck"),
                "reisetage": v.get("reisetage"),
            }
            key = (heute, name, route)
            if key in idx:
                if v["preis"] < idx[key]["preis"]:
                    idx[key].update(rec)
            else:
                idx[key] = rec
                hist.append(rec)
    grenze = (dt.date.today() - dt.timedelta(days=tage - 1)).isoformat()
    hist = [r for r in hist if r["datum"] >= grenze]
    hist_path.write_text(json.dumps(hist, ensure_ascii=False, indent=2), encoding="utf-8")
    return hist


def schreibe_excel(xlsx_path, best_pro_konst, n_best, historie=None):
    historie = historie or []
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEAD  = PatternFill("solid", fgColor="2F5496"); HEADF  = Font(bold=True, color="FFFFFF")
    HIN   = PatternFill("solid", fgColor="1F6B3A"); HINF   = Font(bold=True, color="FFFFFF")
    RUECK = PatternFill("solid", fgColor="7B3F00"); RUECKF = Font(bold=True, color="FFFFFF")
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    RED   = Font(bold=True, color="C00000")           # Direktfluege (0 Umstiege) hervorheben
    ROSE  = PatternFill("solid", fgColor="FFEAF0")     # sehr leichter Rosé-Ton als Hintergrund
    CENTER = Alignment(horizontal="center", vertical="center")
    stand = dt.datetime.now().strftime("%Y-%m-%dT%H:%M")

    def mark_direkt(ws, gruppen):
        """gruppen = Liste von (umstiege_spalte, [zu_faerbende_spalten]).
        Ist die Umstiege-Zelle einer Richtung 0 (Direktflug/Nonstop), werden alle
        Spalten dieser Richtung (Datum, Dauer, Umstiege) rot + sehr leicht rosé hinterlegt."""
        row = ws.max_row
        for umst_col, cols in gruppen:
            if ws.cell(row=row, column=umst_col).value == 0:
                for col in cols:
                    cell = ws.cell(row=row, column=col)
                    cell.font = RED
                    cell.fill = ROSE

    def gruppenheader(ws, meta):
        """meta = Liste von (Gruppenname, Farbe, Schrift, [Spaltennamen]).
           Schreibt Zeile 1 (Gruppenbezeichnungen, zusammengefasst) und Zeile 2 (Spaltennamen)."""
        col = 1
        for gruppe, fill, font, cols in meta:
            n = len(cols)
            if gruppe:
                start = get_column_letter(col)
                end   = get_column_letter(col + n - 1)
                ws.merge_cells(f"{start}1:{end}1")
                c = ws[f"{start}1"]
                c.value = gruppe; c.fill = fill; c.font = font; c.alignment = CENTER
            else:
                for i in range(n):
                    c = ws.cell(row=1, column=col + i)
                    c.fill = fill; c.font = font
            for i, name in enumerate(cols):
                c = ws.cell(row=2, column=col + i)
                c.value = name; c.fill = fill; c.font = font; c.alignment = CENTER
            col += n
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = "A3"

    def autofit(ws):
        for col in ws.columns:
            maxlen = max((len(str(c.value)) for c in col if c.value is not None), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(maxlen + 2, 60)

    link_font = Font(color="0563C1", underline="single")

    def set_link(ws, url):
        c = ws.cell(row=ws.max_row, column=ws.max_column)
        c.value = "Suche oeffnen"; c.hyperlink = url; c.font = link_font

    def fmt(minutes):
        return f"{minutes//60} Std. {minutes%60} Min."

    def datum(iso):
        d = dt.date.fromisoformat(iso)
        return f"{d.day:02d}.{d.month:02d}.{d.year}"

    wb = Workbook()

    # --- Best Of Tab ---
    ws = wb.active; ws.title = "Best Of"
    # Spalten: [Allgemein x6] [Hinflug x3] [Rückflug x3] [Rest x3]
    gruppenheader(ws, [
        (None,       HEAD,  HEADF,  ["Rang", "Reise", "Abflug", "Ziel", "Preis (EUR)", "Reisetage"]),
        ("Hinflug",  HIN,   HINF,   ["Datum", "Dauer", "Umstiege"]),
        ("Rueckflug",RUECK, RUECKF, ["Datum", "Dauer", "Umstiege"]),
        (None,       HEAD,  HEADF,  ["Airline", "Stand", "Link"]),
    ])
    alle = [(name, v) for name, routen in best_pro_konst.items() for v in routen.values()]
    alle.sort(key=lambda x: x[1]["preis"])
    for i, (name, v) in enumerate(alle[:n_best], 1):
        ws.append([i, name, v["orig"], v["dest"], v["preis"], v.get("reisetage"),
                   datum(v["dep"]), fmt(v["dauer_hin_min"]),  v.get("umstiege_hin"),
                   datum(v["ret"]), fmt(v["dauer_rueck_min"]), v.get("umstiege_rueck"),
                   v["airline"], stand, ""])
        set_link(ws, v.get("url") or search_url(v["orig"], v["dest"], v["dep"], v["ret"]))
        if v["unter"]:
            ws.cell(row=ws.max_row, column=5).fill = GREEN
        mark_direkt(ws, [(9, [7, 8, 9]), (12, [10, 11, 12])])  # Hin / Rueck: Datum, Dauer, Umstiege
    autofit(ws)

    # --- Reise-Tabs ---
    for name, routen in best_pro_konst.items():
        ws2 = wb.create_sheet(name[:31])
        gruppenheader(ws2, [
            (None,       HEAD,  HEADF,  ["Abflug", "Ziel", "Preis (EUR)", "Reisetage"]),
            ("Hinflug",  HIN,   HINF,   ["Datum", "Dauer", "Umstiege"]),
            ("Rueckflug",RUECK, RUECKF, ["Datum", "Dauer", "Umstiege"]),
            (None,       HEAD,  HEADF,  ["Airline", "Stand", "Link"]),
        ])
        for v in sorted(routen.values(), key=lambda x: x["preis"]):
            ws2.append([v["orig"], v["dest"], v["preis"], v.get("reisetage"),
                        datum(v["dep"]), fmt(v["dauer_hin_min"]),  v.get("umstiege_hin"),
                        datum(v["ret"]), fmt(v["dauer_rueck_min"]), v.get("umstiege_rueck"),
                        v["airline"], stand, ""])
            set_link(ws2, v.get("url") or search_url(v["orig"], v["dest"], v["dep"], v["ret"]))
            if v["unter"]:
                ws2.cell(row=ws2.max_row, column=3).fill = GREEN
            mark_direkt(ws2, [(7, [5, 6, 7]), (10, [8, 9, 10])])  # Hin / Rueck: Datum, Dauer, Umstiege
        autofit(ws2)
        baue_chart(ws2, name, historie)

    # --- Historie Tab (Rohdaten der letzten 7 Tage) ---
    wsh = wb.create_sheet("Historie")
    hcols = ["Datum", "Reise", "Verbindung", "Abflug", "Ziel", "Preis (EUR)",
             "Hinflug", "Rueckflug", "Umstiege Hin", "Umstiege Rueck", "Reisetage", "Airline"]
    wsh.append(hcols)
    for c in wsh[1]:
        c.font = HEADF; c.fill = HEAD; c.alignment = CENTER
    wsh.freeze_panes = "A2"
    for r in sorted(historie, key=lambda x: (x["konstellation"], x["route"], x["datum"])):
        wsh.append([datum(r["datum"]), r["konstellation"], r["route"], r["orig"], r["dest"],
                    r["preis"], datum(r["dep"]), datum(r["ret"]),
                    r.get("umstiege_hin"), r.get("umstiege_rueck"), r.get("reisetage"),
                    r.get("airline")])
        mark_direkt(wsh, [(9, [7, 9]), (10, [8, 10])])  # Hin / Rueck: Datum, Umstiege (keine Dauer-Spalte)
    autofit(wsh)

    try:
        wb.save(xlsx_path)
    except PermissionError:
        print(f"[!] {xlsx_path.name} ist gesperrt (in Excel geoeffnet?) - Excel uebersprungen.")


# Mobile Ansicht: eine selbst enthaltene, statische HTML-Datei OHNE JavaScript.
# Alle Inhalte werden beim Lauf fertig gerendert; Navigation ueber Anker-Links und
# <details>-Abschnitte. Funktioniert damit auch in Viewern, die kein JavaScript
# ausfuehren (iOS-Dateivorschau / Quick Look / OneDrive-Vorschau).
_HTML_KOPF = """<!doctype html>
<html lang="de">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Urlaubspreise</title>
<style>
:root{--bg:#f5f4f0;--card:#ffffff;--text:#1a1a18;--mut:#6b6a64;--line:#e2e0d8;--acc:#0b5fa5;--ok-bg:#dcefdc;--ok-tx:#1f6b3a;--warn:#c00000;}
@media(prefers-color-scheme:dark){:root{--bg:#161614;--card:#232320;--text:#ececea;--mut:#a3a29b;--line:#3a3936;--acc:#7db8e8;--ok-bg:#1f3a26;--ok-tx:#8fd3a0;--warn:#ff8a80;}}
*{box-sizing:border-box}
body{margin:0;background:var(--bg);color:var(--text);font-family:-apple-system,"Segoe UI",Roboto,sans-serif;font-size:15px;}
@media (max-device-width:500px) and (min-width:600px){body{zoom:2.4;}}
@media (min-device-width:501px) and (max-device-width:900px) and (min-width:960px){body{zoom:1.5;}}
header{position:sticky;top:0;background:var(--card);border-bottom:1px solid var(--line);padding:8px 14px 0;z-index:2;}
h1{font-size:17px;margin:0;font-weight:600;}
.stand{font-size:12px;color:var(--mut);margin:2px 0 6px;}
.topnav{display:flex;gap:20px;}
.topnav a{color:var(--acc);text-decoration:none;font-size:14px;font-weight:600;padding:2px 0 9px;}
main{padding:10px 10px 30px;max-width:720px;margin:0 auto;}
section{scroll-margin-top:92px;}
h2{font-size:16px;margin:18px 2px 8px;}
.sub{font-size:12.5px;color:var(--mut);margin:0 2px 8px;}
.card{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:10px 12px;margin:0 0 8px;display:flex;justify-content:space-between;gap:10px;align-items:center;}
.card .t{font-weight:600;font-size:15px;margin:0;}
.card .s{font-size:12.5px;color:var(--mut);margin:2px 0 0;}
.preis{font-size:17px;font-weight:700;white-space:nowrap;text-align:right;}
.badge{display:inline-block;background:var(--ok-bg);color:var(--ok-tx);font-size:12px;border-radius:6px;padding:1px 7px;margin-top:4px;}
.direkt{color:var(--warn);font-weight:600;}
a.link{color:var(--acc);font-size:12.5px;}
details{background:var(--card);border:1px solid var(--line);border-radius:12px;margin:0 0 8px;padding:0 10px;}
details[open]{padding-bottom:8px;}
summary{display:flex;justify-content:space-between;align-items:center;gap:10px;padding:10px 2px;cursor:pointer;list-style:none;}
summary::-webkit-details-marker{display:none}
summary .t{font-weight:600;font-size:15px;margin:0;}
summary .s{font-size:12.5px;color:var(--mut);margin:2px 0 0;}
details .card{background:var(--bg);}
</style>
</head>
<body>
"""


def schreibe_html(html_path, best_pro_konst, n_best, historie=None, schwellen=None):
    """Schreibt die mobile HTML-Ansicht: statische, selbst enthaltene Seite ohne
    JavaScript mit denselben Daten wie die Excel (Best Of, Reisen, 7-Tage-Historie)."""
    from html import escape
    historie = historie or []
    schwellen = schwellen or {}

    def eur(p):
        return f"{p:,.0f}".replace(",", ".") + " €"

    def fd(iso):
        return f"{iso[8:10]}.{iso[5:7]}." if iso else ""

    def fd_lang(iso):
        return f"{iso[8:10]}.{iso[5:7]}.{iso[0:4]}"

    def fdau(mi):
        return "" if mi is None else f"{mi // 60}h{mi % 60:02d}"

    def umst(u):
        if u == 0:
            return '<span class="direkt">direkt</span>'
        if u is None:
            return "?"
        return f"{u} Stopp" + ("s" if u > 1 else "")

    def karte(v, konst=None, link=False):
        t = f"{escape(v['orig'])} → {escape(v['dest'])}" + (f" · {escape(konst)}" if konst else "")
        zeilen = [f"{fd(v['dep'])}–{fd(v['ret'])}" + (f" · {v['reisetage']} Tage" if v.get("reisetage") else "")]
        if v.get("etappen"):
            zeilen.extend(f"{escape(e['von'])} → {escape(e['nach'])} {fd(e['datum'])} · "
                          f"{eur(e['preis'])} ({umst(e['umstiege'])})" for e in v["etappen"])
        elif v.get("dauer_hin_min") is not None:
            zeilen.append(f"Hin {fdau(v['dauer_hin_min'])} ({umst(v.get('umstiege_hin'))}) · "
                          f"Rück {fdau(v.get('dauer_rueck_min'))} ({umst(v.get('umstiege_rueck'))})")
        else:
            zeilen.append(f"Hin {umst(v.get('umstiege_hin'))} · Rück {umst(v.get('umstiege_rueck'))}")
        rest = escape(v.get("airline") or "")
        if link:
            url = v.get("url") or search_url(v["orig"], v["dest"], v["dep"], v["ret"])
            rest += (" · " if rest else "") + f'<a class="link" href="{escape(url)}" target="_blank" rel="noopener">Suche öffnen</a>'
        if rest:
            zeilen.append(rest)
        subs = "".join(f'<p class="s">{z}</p>' for z in zeilen)
        badge = '<span class="badge">unter Schwelle</span>' if v.get("unter") else ""
        return (f'<div class="card"><div><p class="t">{t}</p>{subs}{badge}</div>'
                f'<div class="preis">{eur(v["preis"])}</div></div>')

    stand = dt.datetime.now().strftime("%d.%m.%Y %H:%M")
    konst_liste = [(name, sorted(routen.values(), key=lambda x: x["preis"]))
                   for name, routen in best_pro_konst.items()]
    konst_liste.sort(key=lambda x: x[1][0]["preis"] if x[1] else 9e9)
    alle = [(name, v) for name, rr in konst_liste for v in rr]
    alle.sort(key=lambda x: x[1]["preis"])

    teile = [_HTML_KOPF,
             f'<header><h1>Urlaubspreise</h1><p class="stand">Stand {stand}</p>'
             '<nav class="topnav"><a href="#bestof">Best Of</a><a href="#reisen">Reisen</a>'
             '<a href="#historie">Historie</a></nav></header><main>']

    teile.append(f'<section id="bestof"><h2>Best Of</h2>'
                 f'<p class="sub">Die {min(n_best, len(alle))} günstigsten Angebote über alle Reisen</p>')
    for name, v in alle[:n_best]:
        teile.append(karte(v, konst=name, link=True))
    if not alle:
        teile.append('<p class="sub">Noch keine Daten.</p>')
    teile.append("</section>")

    teile.append('<section id="reisen"><h2>Reisen</h2>'
                 '<p class="sub">Antippen zum Aufklappen – beste Verbindung je Route</p>')
    for name, rr in konst_liste:
        s = schwellen.get(name)
        sub = f"{len(rr)} Routen" + (f" · Schwelle {eur(s)}" if s else "")
        ab = f"ab {eur(rr[0]['preis'])}" if rr else "–"
        teile.append(f'<details><summary><span><p class="t">{escape(name)}</p><p class="s">{sub}</p></span>'
                     f'<span class="preis">{ab}</span></summary>')
        teile.extend(karte(v, link=True) for v in rr)
        teile.append("</details>")
    teile.append("</section>")

    teile.append('<section id="historie"><h2>Historie</h2>'
                 '<p class="sub">Tagesbestpreise der letzten 7 Tage</p>')
    tage = sorted({r["datum"] for r in historie}, reverse=True)
    for i, tag in enumerate(tage):
        rows = sorted((r for r in historie if r["datum"] == tag), key=lambda r: r["preis"])
        teile.append(f'<details{" open" if i == 0 else ""}><summary><span>'
                     f'<p class="t">{fd_lang(tag)}</p><p class="s">{len(rows)} Einträge</p></span></summary>')
        teile.extend(karte(r, konst=r.get("konstellation")) for r in rows)
        teile.append("</details>")
    if not tage:
        teile.append('<p class="sub">Noch keine Historie.</p>')
    teile.append("</section></main></body></html>")

    try:
        html_path.write_text("\n".join(teile), encoding="utf-8")
    except PermissionError:
        print(f"[!] {html_path.name} ist gesperrt - HTML uebersprungen.")


# App-Ansicht: interaktive Ein-Datei-Anwendung (Tabs, Filter) fuer echte Browser.
# Wird als best-of-app.html geschrieben und nach jedem Lauf als GitHub-Pages-Seite
# (index.html) veroeffentlicht, damit sie auf dem Handy unter fester Adresse laeuft.
_APP_VORLAGE = """<!doctype html>
<html lang="de">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="default">
<meta name="theme-color" content="#f5f4f0">
<title>Urlaubspreise</title>
<style>
:root{--bg:#f5f4f0;--card:#ffffff;--text:#1a1a18;--mut:#6b6a64;--line:#e2e0d8;--acc:#0b5fa5;--ok-bg:#dcefdc;--ok-tx:#1f6b3a;--warn:#c00000;}
@media(prefers-color-scheme:dark){:root{--bg:#161614;--card:#232320;--text:#ececea;--mut:#a3a29b;--line:#3a3936;--acc:#7db8e8;--ok-bg:#1f3a26;--ok-tx:#8fd3a0;--warn:#ff8a80;}}
*{box-sizing:border-box}
body{margin:0;background:var(--bg);color:var(--text);font-family:-apple-system,"Segoe UI",Roboto,sans-serif;font-size:15px;}
header{position:sticky;top:0;background:var(--card);border-bottom:1px solid var(--line);padding:10px 14px;z-index:2;display:flex;align-items:center;gap:6px;}
header h1{font-size:17px;margin:0;font-weight:600;}
header .stand{font-size:12px;color:var(--mut);margin:2px 0 0;}
main{padding:10px 10px 80px;max-width:720px;margin:0 auto;}
.card{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:10px 12px;margin:0 0 8px;display:flex;justify-content:space-between;gap:10px;align-items:center;}
.card .t{font-weight:600;font-size:15px;margin:0;}
.card .s{font-size:12.5px;color:var(--mut);margin:2px 0 0;}
.preis{font-size:17px;font-weight:700;white-space:nowrap;text-align:right;}
.badge{display:inline-block;background:var(--ok-bg);color:var(--ok-tx);font-size:12px;border-radius:6px;padding:1px 7px;margin-top:4px;}
.direkt{color:var(--warn);font-weight:600;}
a.link{color:var(--acc);font-size:12.5px;}
nav{position:fixed;bottom:0;left:0;right:0;background:var(--card);border-top:1px solid var(--line);display:flex;z-index:2;padding-bottom:env(safe-area-inset-bottom);}
nav a{flex:1;text-align:center;padding:7px 0 9px;color:var(--mut);text-decoration:none;font-size:12px;}
nav a.on{color:var(--acc);font-weight:600;}
nav .ic{display:block;font-size:19px;line-height:1.25;}
.chips{display:flex;gap:6px;overflow-x:auto;padding:2px 0 8px;}
.chip{border:1px solid var(--line);background:var(--card);color:var(--mut);border-radius:999px;padding:4px 12px;font-size:13px;white-space:nowrap;}
.chip.on{background:var(--acc);border-color:var(--acc);color:#fff;}
.tag{font-size:12px;color:var(--mut);margin:14px 2px 6px;font-weight:600;}
.back{background:none;border:none;font-size:26px;color:var(--acc);padding:0 8px 0 0;cursor:pointer;line-height:1;}
.leer{color:var(--mut);text-align:center;padding:30px 0;}
.seg{margin-left:auto;display:flex;border:1px solid var(--line);border-radius:999px;overflow:hidden;flex-shrink:0;}
.seg button{background:none;border:none;font-size:13px;padding:5px 11px;color:var(--mut);cursor:pointer;}
.seg button.on{background:var(--acc);color:#fff;}
</style>
</head>
<body>
<header id="hdr"></header>
<main id="main"></main>
<nav id="nav"></nav>
<script>
const D = __DATEN__;
D.konstellationen.forEach(k => k.routen.forEach(r => r.schwelle = k.schwelle));
{ const s = {}; D.konstellationen.forEach(k => s[k.name] = k.schwelle);
  D.historie.forEach(r => r.schwelle = (r.konstellation in s) ? s[r.konstellation] : null); }
const M = document.getElementById("main"), H = document.getElementById("hdr"), N = document.getElementById("nav");
const esc = s => String(s == null ? "" : s).replace(/[&<>"]/g, c => ({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"}[c]));
const eur = p => Number(p).toLocaleString("de-DE") + " €";
const fd = iso => iso ? iso.slice(8,10) + "." + iso.slice(5,7) + "." : "";
const fdLang = iso => iso.slice(8,10) + "." + iso.slice(5,7) + "." + iso.slice(0,4);
const fdau = mi => mi == null ? "" : Math.floor(mi/60) + "h" + String(mi%60).padStart(2,"0");
const umst = u => u === 0 ? '<span class="direkt">direkt</span>' : (u == null ? "?" : u + " Stopp" + (u > 1 ? "s" : ""));
const alle = D.konstellationen.flatMap(k => k.routen.map(r => Object.assign({konst: k.name}, r)));
const NAMEN = [...new Set(D.historie.map(r => r.konstellation))];
let hFilter = "";
let modus = "abs";
try{ modus = localStorage.getItem("preismodus") || "abs"; }catch(e){}

function setModus(m){ modus = m; try{ localStorage.setItem("preismodus", m); }catch(e){} render(); }
function sortKey(v){ return modus === "rel" ? (v.schwelle != null ? v.preis - v.schwelle : 1e12) : v.preis; }
function relTxt(d){ return (d > 0 ? "+" : "") + Number(d).toLocaleString("de-DE") + " €"; }

function preisHtml(v){
  if(modus === "rel" && v.schwelle != null){
    const d = v.preis - v.schwelle;
    return '<div class="preis" style="color:' + (d <= 0 ? "var(--ok-tx)" : "var(--warn)") + '">' + relTxt(d) +
      '<div style="font-size:12px;color:var(--mut);font-weight:400;">' + eur(v.preis) + "</div></div>";
  }
  return '<div class="preis">' + eur(v.preis) + "</div>";
}

function kopf(titel, sub, zurueck){
  const seg = '<div class="seg">' +
    '<button' + (modus === "abs" ? ' class="on"' : "") + ' onclick="setModus(\\'abs\\')" title="Absolute Preise">&euro;</button>' +
    '<button' + (modus === "rel" ? ' class="on"' : "") + ' onclick="setModus(\\'rel\\')" title="Relativ zur Schwelle">&plusmn;</button></div>';
  H.innerHTML = (zurueck ? '<button class="back" onclick="history.back()">&#8249;</button>' : "") +
    '<div style="flex:1;min-width:0;"><h1>' + esc(titel) + "</h1>" + (sub ? '<p class="stand">' + esc(sub) + "</p>" : "") + "</div>" + seg;
}

function karte(v, mitKonst){
  return '<div class="card"><div>' +
    '<p class="t">' + esc(v.orig) + " &#8594; " + esc(v.dest) + (mitKonst ? " &middot; " + esc(v.konst) : "") + "</p>" +
    '<p class="s">' + fd(v.dep) + "&ndash;" + fd(v.ret) + (v.reisetage ? " &middot; " + v.reisetage + " Tage" : "") + "</p>" +
    (v.etappen ? v.etappen.map(e => '<p class="s">' + esc(e.von) + " &#8594; " + esc(e.nach) + " " + fd(e.datum) + " &middot; " + eur(e.preis) + " (" + umst(e.umstiege) + ")</p>").join("") :
      '<p class="s">Hin ' + fdau(v.dauer_hin_min) + " (" + umst(v.umstiege_hin) + ") &middot; R&uuml;ck " + fdau(v.dauer_rueck_min) + " (" + umst(v.umstiege_rueck) + ")</p>") +
    '<p class="s">' + esc(v.airline || "") + (v.url ? ' &middot; <a class="link" href="' + esc(v.url) + '" target="_blank" rel="noopener">Suche &ouml;ffnen</a>' : "") + "</p>" +
    (v.unter ? '<span class="badge">unter Schwelle</span>' : "") +
    "</div>" + preisHtml(v) + "</div>";
}

function vBestof(){
  kopf("Urlaubspreise", "Stand " + D.stand + (modus === "rel" ? " · sortiert relativ zur Schwelle" : " · Top " + D.n_best));
  const top = alle.slice().sort((a,b) => sortKey(a) - sortKey(b)).slice(0, D.n_best);
  M.innerHTML = top.length ? top.map(v => karte(v, true)).join("") : '<p class="leer">Noch keine Daten.</p>';
}

function bestKey(k){ return k.routen.length ? sortKey(k.routen[0]) : 1e12; }

function vReisen(){
  kopf("Reisen", D.konstellationen.length + " Reisen");
  const idx = D.konstellationen.map((k, i) => i);
  if(modus === "rel") idx.sort((a,b) => bestKey(D.konstellationen[a]) - bestKey(D.konstellationen[b]));
  M.innerHTML = idx.map(i => {
    const k = D.konstellationen[i];
    const best = k.routen.length ? k.routen[0] : null;
    let ab = "&ndash;";
    if(best){
      if(modus === "rel" && best.schwelle != null){
        ab = '<span style="color:' + (best.preis - best.schwelle <= 0 ? "var(--ok-tx)" : "var(--warn)") + '">ab ' + relTxt(best.preis - best.schwelle) + "</span>";
      } else {
        ab = "ab " + eur(best.preis);
      }
    }
    return '<div class="card" style="cursor:pointer" onclick="location.hash=\\'#reise/' + i + '\\'">' +
      '<div><p class="t">' + esc(k.name) + '</p><p class="s">' + k.routen.length + " Routen" + (k.schwelle ? " &middot; Schwelle " + eur(k.schwelle) : "") + "</p></div>" +
      '<div class="preis">' + ab + ' <span style="color:var(--mut)">&#8250;</span></div></div>';
  }).join("") || '<p class="leer">Noch keine Daten.</p>';
}

function vReise(i){
  const k = D.konstellationen[i];
  if(!k){ location.hash = "#reisen"; return; }
  kopf(k.name, (k.schwelle ? "Schwelle " + eur(k.schwelle) + " · " : "") + "beste Verbindung je Route", true);
  M.innerHTML = k.routen.map(v => karte(v, false)).join("") || '<p class="leer">Noch keine Daten.</p>';
}

function setF(i){ hFilter = i < 0 ? "" : NAMEN[i]; vHistorie(); }

function histKarte(r){
  return '<div class="card"><div>' +
    '<p class="t">' + esc(r.orig) + " &#8594; " + esc(r.dest) + (hFilter ? "" : " &middot; " + esc(r.konstellation)) + "</p>" +
    '<p class="s">' + fd(r.dep) + "&ndash;" + fd(r.ret) + " &middot; Hin " + umst(r.umstiege_hin) + " &middot; R&uuml;ck " + umst(r.umstiege_rueck) + "</p>" +
    '<p class="s">' + esc(r.airline || "") + "</p>" +
    "</div>" + preisHtml(r) + "</div>";
}

function vHistorie(){
  kopf("Historie", "Tagesbestpreise der letzten 7 Tage");
  const rows = D.historie.filter(r => !hFilter || r.konstellation === hFilter);
  const tage = [...new Set(rows.map(r => r.datum))].sort().reverse();
  let html = '<div class="chips"><button class="chip' + (hFilter ? "" : " on") + '" onclick="setF(-1)">Alle</button>' +
    NAMEN.map((n, i) => '<button class="chip' + (hFilter === n ? " on" : "") + '" onclick="setF(' + i + ')">' + esc(n) + "</button>").join("") + "</div>";
  html += tage.map(t => '<p class="tag">' + fdLang(t) + "</p>" +
    rows.filter(r => r.datum === t).sort((a,b) => sortKey(a) - sortKey(b)).map(histKarte).join("")).join("");
  M.innerHTML = tage.length ? html : html + '<p class="leer">Noch keine Historie.</p>';
}

function zeichneNav(akt){
  N.innerHTML = [["#bestof","&#127942;","Best Of"],["#reisen","&#9992;&#65039;","Reisen"],["#historie","&#128197;","Historie"]]
    .map(x => '<a href="' + x[0] + '" class="' + (akt === x[0].slice(1) ? "on" : "") + '"><span class="ic">' + x[1] + "</span>" + x[2] + "</a>").join("");
}

function render(){
  const h = (location.hash || "#bestof").slice(1);
  if(h.indexOf("reise/") === 0){ zeichneNav("reisen"); vReise(parseInt(h.slice(6), 10)); }
  else if(h === "reisen"){ zeichneNav("reisen"); vReisen(); }
  else if(h === "historie"){ zeichneNav("historie"); vHistorie(); }
  else { zeichneNav("bestof"); vBestof(); }
  window.scrollTo(0, 0);
}
window.addEventListener("hashchange", render);
render();
</script>
</body>
</html>
"""


def schreibe_app_html(app_path, best_pro_konst, n_best, historie=None, schwellen=None):
    """Schreibt die interaktive App-Ansicht (Tabs/Filter, braucht JavaScript) als
    eigene Datei; wird zusaetzlich als GitHub-Pages-Seite veroeffentlicht."""
    historie = historie or []
    schwellen = schwellen or {}
    stand = dt.datetime.now().strftime("%d.%m.%Y %H:%M")
    konst = []
    for name, routen in best_pro_konst.items():
        rr = []
        for v in sorted(routen.values(), key=lambda x: x["preis"]):
            r = dict(v)
            r["url"] = v.get("url") or search_url(v["orig"], v["dest"], v["dep"], v["ret"])
            rr.append(r)
        konst.append({"name": name, "schwelle": schwellen.get(name), "routen": rr})
    konst.sort(key=lambda k: k["routen"][0]["preis"] if k["routen"] else 9e9)
    payload = {"stand": stand, "n_best": n_best, "konstellationen": konst, "historie": historie}
    daten = json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")
    try:
        app_path.write_text(_APP_VORLAGE.replace("__DATEN__", daten), encoding="utf-8")
    except PermissionError:
        print(f"[!] {app_path.name} ist gesperrt - App-HTML uebersprungen.")


def veroeffentliche_html(pfad, repo):
    """Laedt die App-HTML als index.html in das GitHub-Repo hoch (GitHub Pages).
    Fehler (kein Netz, gh fehlt, ...) brechen den Lauf nicht ab."""
    import subprocess
    try:
        inhalt = base64.b64encode(pfad.read_bytes()).decode()
        body = {"message": "Stand " + dt.datetime.now().strftime("%Y-%m-%d %H:%M"), "content": inhalt}
        # CREATE_NO_WINDOW: sonst blitzt bei jedem Lauf kurz ein Konsolenfenster auf
        kein_fenster = subprocess.CREATE_NO_WINDOW
        r = subprocess.run(["gh", "api", f"repos/{repo}/contents/index.html", "-q", ".sha"],
                           capture_output=True, text=True, timeout=60, creationflags=kein_fenster)
        if r.returncode == 0 and r.stdout.strip():
            body["sha"] = r.stdout.strip()
        vorher = subprocess.run(["gh", "api", f"repos/{repo}/pages/builds/latest", "-q", ".status"],
                                capture_output=True, text=True, timeout=60, creationflags=kein_fenster)
        r = subprocess.run(["gh", "api", "-X", "PUT", f"repos/{repo}/contents/index.html", "--input", "-"],
                           input=json.dumps(body), capture_output=True, text=True, timeout=120,
                           creationflags=kein_fenster)
        if r.returncode != 0 and '"sha"' in r.stderr:
            # sha-Abfrage war fehlgeschlagen/leer (transient): einmal frisch holen und wiederholen
            s2 = subprocess.run(["gh", "api", f"repos/{repo}/contents/index.html", "-q", ".sha"],
                                capture_output=True, text=True, timeout=60, creationflags=kein_fenster)
            if s2.returncode == 0 and s2.stdout.strip():
                body["sha"] = s2.stdout.strip()
                r = subprocess.run(["gh", "api", "-X", "PUT", f"repos/{repo}/contents/index.html", "--input", "-"],
                                   input=json.dumps(body), capture_output=True, text=True, timeout=120,
                                   creationflags=kein_fenster)
        if r.returncode != 0:
            print(f"[!] Veroeffentlichung fehlgeschlagen: {r.stderr.strip()[:200]}")
        elif vorher.returncode == 0 and vorher.stdout.strip() == "errored":
            # Voriger Pages-Build war kaputt (GitHub-Infrastruktur): zusaetzlich zum
            # automatischen Build des neuen Commits explizit einen Neuaufbau anstossen.
            subprocess.run(["gh", "api", "-X", "POST", f"repos/{repo}/pages/builds"],
                           capture_output=True, text=True, timeout=60, creationflags=kein_fenster)
            print("Hinweis: voriger Pages-Build fehlgeschlagen - Neuaufbau angestossen.")
    except Exception as e:
        print(f"[!] Veroeffentlichung fehlgeschlagen: {str(e)[:200]}")


def baue_chart(ws, name, historie):
    """Fuegt unter der Tabelle eine Preis-Pivot (Datum x Verbindung) plus Liniendiagramm ein."""
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.marker import Marker
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    recs = [r for r in historie if r["konstellation"] == name]
    if not recs:
        return
    dates = sorted({r["datum"] for r in recs})
    routes = sorted({r["route"] for r in recs})
    preis = {}
    for r in recs:
        k = (r["datum"], r["route"])
        if k not in preis or r["preis"] < preis[k]:
            preis[k] = r["preis"]

    pstart = ws.max_row + 3
    bold = Font(bold=True)
    c = ws.cell(row=pstart, column=1, value="Datum"); c.font = bold
    for j, rt in enumerate(routes, start=2):
        c = ws.cell(row=pstart, column=j, value=rt); c.font = bold
    for i, d in enumerate(dates, start=1):
        ws.cell(row=pstart + i, column=1, value=d[8:10] + "." + d[5:7] + "." + d[0:4])
        for j, rt in enumerate(routes, start=2):
            val = preis.get((d, rt))
            if val is not None:
                ws.cell(row=pstart + i, column=j, value=val)

    chart = LineChart()
    chart.title = f"Preis-Historie {name} (7 Tage)"
    chart.y_axis.title = "Preis (EUR)"
    chart.x_axis.title = "Datum"
    chart.height = 8; chart.width = 18
    data = Reference(ws, min_col=2, max_col=1 + len(routes), min_row=pstart, max_row=pstart + len(dates))
    cats = Reference(ws, min_col=1, min_row=pstart + 1, max_row=pstart + len(dates))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    # Markierungspunkte: macht auch einen einzelnen Tag (1 Datenpunkt) sichtbar -
    # ein Liniendiagramm zeichnet sonst erst ab 2 Punkten eine Linie.
    for s in chart.series:
        s.marker = Marker(symbol="circle", size=7)
        s.smooth = False

    # Achsenbeschriftung erzwingen (openpyxl blendet sie sonst per Default aus) -
    # so sind sowohl Preis- als auch Datums-Skala sichtbar.
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.y_axis.numFmt = "#,##0"
    chart.y_axis.majorGridlines = ChartLines()
    # Preis-Skala an die Daten anpassen -> auch bei nur 1 Tag gut ablesbar.
    vals = list(preis.values())
    if vals:
        lo, hi = min(vals), max(vals)
        spanne = hi - lo
        pad = max(50, int(spanne * 0.15)) if spanne else max(50, int(hi * 0.05))
        chart.y_axis.scaling.min = max(0, lo - pad)
        chart.y_axis.scaling.max = hi + pad

    anchor = get_column_letter(1 + len(routes) + 2)
    ws.add_chart(chart, f"{anchor}{pstart}")


if __name__ == "__main__":
    main()
